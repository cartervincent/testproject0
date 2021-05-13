# -*- coding: UTF-8 -*-

import time

shijian = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
print(shijian)

x = 5.123
a = int(x)
print(oct(a))

import openpyxl

# 用openpyxl模块打开Excel文档
wb = openpyxl.load_workbook(
    'D:\\Kelvin\\运维\\日报\\域名状况监测\\01-02-03域名检测 - 2020.09.02\\01域名检测 - 2020.09.02\\01域名检测 - 2020.09.02.xlsx')
rlist = openpyxl.load_workbook('D:\\Kelvin\\运维\\日报\\CDN例行性检测\\CDN例行性检测_2020.09.04.xlsx')
print('文件类型是:', type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>

excel_name = []
excel_name.append(wb.sheetnames)
# 从工作簿中取得工作表
#
print(wb.sheetnames)  # ['Sheet1', 'Sheet2', 'Sheet3']
print(rlist.sheetnames)

'''sheet = wb['Sheet1']
print(sheet)  # <Worksheet "Sheet1"> 通过表名取得表格
anotherSheet = wb.active
print(anotherSheet)  # <Worksheet "Sheet3"> 活动表，即工作薄在Excel打开时出现的工作表'''

sheet1 = rlist['SG-PERFORMANCE-2']
print(sheet1)
anothersheet = rlist.active
print('active page：', anothersheet, '即工作薄在Excel打开时出现的工作表.')

print(excel_name)

ws = wb.active
print(ws)

for ws in wb:
    print(ws.title)
aa = ws.active_cell
print(aa)
